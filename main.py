# Company name: B10
# TAX INVOICE: B9
# Date: C18
# Order quantity: E22

# Steps:
# Get all sheets
# For each sheet check for company name in B10
# Check dominant date
# Get date from C18 (ignore if not matching dominant date)
# Get order amount
# Add to df
#
# Output:
# Output DF to xls

import re
import sys
from collections import Counter
from pathlib import Path

import pandas as pd
from dateutil import parser
from openpyxl import load_workbook
from pprint import pprint

PATH = Path(r"D:\Desktop\Invoices")
OUTPUT_PATH = Path(r'D:/Desktop/output.xlsx')

output_df = pd.DataFrame()

output_dict = {}


def main():
    print(f'Accessing directory {PATH} ...')
    for file in PATH.glob('*.xls*'):

        # Skip placeholder files
        if file.name.startswith('~'):
            continue

        print(f'Processing file \'{file.name}\'...')
        try:
            process_file(file)
        except Exception as e:
            print(f'Error while processing {file.name}: {e}', file=sys.stderr)

    # Format dictionary suitable for output to excel
    df = pd.DataFrame.from_dict(output_dict, orient='index')
    df.sort_index(0, inplace=True)
    df.sort_index(1, inplace=True)
    df.rename(mapper=lambda x: x.strftime('%b, %Y'), axis=1, inplace=True)
    df.to_excel(OUTPUT_PATH, engine='openpyxl')
    print(f'Output file written to {OUTPUT_PATH}')




def process_file(file):
    wb = load_workbook(file)
    dominant_date = get_dominant_date(wb)

    for sheetname in wb.sheetnames:

        try:
            sheet = wb[sheetname]
            company_name = get_company_name(sheet)
            date = get_date(sheet)
            order_amount = get_order_amount(sheet)

        except Exception as e:
            print(f'Skipping sheet {sheetname} in {file.name}: {e}', file=sys.stderr)

        else:
            # Ignore sheet if date is not dominant date
            if date != dominant_date:
                continue

            # Only update main dictionary if sheet was parsed successfully
            company = output_dict.setdefault(company_name, {})
            company[date] = order_amount


def get_dominant_date(wb):
    date_counter = Counter()

    for sheetname in wb.sheetnames:
        sheet = wb[sheetname]

        # Only count dates from valid sheets
        if validate_sheet(sheet):
            date = get_date(sheet)
            date_counter[date] += 1

    return date_counter.most_common(1)[0][0]


def validate_sheet(sheet):
    try:
        get_company_name(sheet)
        get_order_amount(sheet)
        get_date(sheet)
    except:
        return False
    else:
        return True


def get_company_name(sheet):
    return sheet['B10'].value


def get_date(sheet):
    raw_string = sheet['C18'].value
    clean_string = re.findall('(\\w+\\s+\\w+).*', raw_string)[0]
    date = parser.parse(clean_string)
    return date


def get_order_amount(sheet):
    raw_string = sheet['E22'].value
    order_amount = re.findall('\\b([\\d,]+)\\b', raw_string)[0]

    # remove commas, if present
    return int(order_amount.replace(',', ''))


if __name__ == '__main__':
    main()
